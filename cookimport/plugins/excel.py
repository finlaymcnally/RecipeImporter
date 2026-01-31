from __future__ import annotations

import datetime as dt
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Callable, Iterable

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

from cookimport import __version__
from cookimport.core.models import (
    ConversionReport,
    ConversionResult,
    MappingConfig,
    RawArtifact,
    RecipeCandidate,
    SheetInspection,
    SheetMapping,
    SkippedRow,
    WorkbookInspection,
)
from cookimport.core.reporting import compute_file_hash
from cookimport.core.scoring import score_recipe_candidate
from cookimport.parsing.tips import (
    extract_tip_candidates_from_candidate,
    partition_tip_candidates,
)
from cookimport.plugins import registry

_HEADER_ALIASES = {
    "name": ["name", "title", "recipe", "recipe name"],
    "ingredients": ["ingredients", "ingredient", "ingredient list", "recipe/ingredients"],
    "instructions": ["instructions", "instruction", "steps", "method", "directions"],
    "description": ["description", "notes", "headnote"],
    "recipeYield": ["yield", "servings", "serves"],
    "sourceUrl": ["url", "source", "link"],
    "tags": ["tags", "keywords", "categories", "cuisine", "type", "tool"],
}

_SECTION_INGREDIENTS = ("ingredient", "ingredients", "ing")
_SECTION_INSTRUCTIONS = ("instruction", "instructions", "direction", "directions", "method", "step", "steps")
_SECTION_NOTES = ("note", "notes", "tip", "tips")

# Regex for detecting section headers in text blobs (e.g., "Ingredients:", "INSTRUCTIONS", "Notes")
_SECTION_HEADER_RE = re.compile(
    r"^\s*(ingredients?|instructions?|directions?|method|steps?|preparation|notes?|tips?)\s*:?\s*$",
    re.IGNORECASE,
)

_ALIAS_LOOKUP: dict[str, str] = {}
_SLUG_RE = re.compile(r"[^a-z0-9]+")


@dataclass
class SheetAnalysis:
    layout: str
    header_row: int | None
    headers: list[str]
    column_aliases: dict[str, list[str]]
    template_cells: dict[str, str]
    tall_keys: dict[str, str]
    confidence: float
    low_confidence: bool
    warnings: list[str]


class ExcelImporter:
    name = "excel"

    def detect(self, path: Path) -> float:
        if path.suffix.lower() in {".xlsx", ".xlsm"}:
            return 0.95
        return 0.0

    def inspect(self, path: Path) -> WorkbookInspection:
        wb_values = load_workbook(path, read_only=True, data_only=True)
        wb_meta = load_workbook(path, read_only=False, data_only=False)
        try:
            sheet_inspections: list[SheetInspection] = []
            sheet_mappings: list[SheetMapping] = []
            for name in wb_values.sheetnames:
                values_sheet = wb_values[name]
                meta_sheet = wb_meta[name]
                merged_map = _merged_cell_map(meta_sheet, values_sheet)
                analysis = _analyze_sheet(values_sheet, meta_sheet, merged_map)
                sheet_inspections.append(
                    SheetInspection(
                        name=name,
                        layout=analysis.layout,
                        header_row=analysis.header_row,
                        inferred_mapping=SheetMapping(
                            sheetName=name,
                            layout=analysis.layout,
                            headerRow=analysis.header_row,
                            lowConfidence=analysis.low_confidence,
                            columnAliases=analysis.column_aliases,
                            templateCells=analysis.template_cells,
                            tallKeys=analysis.tall_keys,
                        ),
                        confidence=analysis.confidence,
                        lowConfidence=analysis.low_confidence,
                        warnings=analysis.warnings,
                    )
                )
                sheet_mappings.append(
                    SheetMapping(
                        sheetName=name,
                        layout=analysis.layout,
                        headerRow=analysis.header_row,
                        lowConfidence=analysis.low_confidence,
                        columnAliases=analysis.column_aliases,
                        templateCells=analysis.template_cells,
                        tallKeys=analysis.tall_keys,
                    )
                )
            mapping_stub = MappingConfig(sheets=sheet_mappings)
            return WorkbookInspection(path=str(path), sheets=sheet_inspections, mappingStub=mapping_stub)
        finally:
            wb_values.close()
            wb_meta.close()

    def convert(
        self,
        path: Path,
        mapping: MappingConfig | None,
        progress_callback: Callable[[str], None] | None = None,
    ) -> ConversionResult:
        if mapping is None:
            if progress_callback:
                progress_callback("Inspecting workbook...")
            inspection = self.inspect(path)
            mapping = inspection.mapping_stub
            inferred = {sheet.name: sheet for sheet in inspection.sheets}
        else:
            if progress_callback:
                progress_callback("Inspecting workbook...")
            inspection = self.inspect(path)
            inferred = {sheet.name: sheet for sheet in inspection.sheets}

        wb_values = load_workbook(path, read_only=True, data_only=True)
        wb_meta = load_workbook(path, read_only=False, data_only=False)
        report = ConversionReport(mappingUsed=mapping)
        recipes: list[RecipeCandidate] = []
        raw_artifacts: list[RawArtifact] = []
        file_hash = compute_file_hash(path)
        overrides = mapping.parsing_overrides if mapping else None
        try:
            for name in wb_values.sheetnames:
                if progress_callback:
                    progress_callback(f"Processing sheet '{name}'...")
                values_sheet = wb_values[name]
                meta_sheet = wb_meta[name]
                merged_map = _merged_cell_map(meta_sheet, values_sheet)
                sheet_mapping = _resolve_sheet_mapping(mapping, name) if mapping else None
                inferred_sheet = inferred.get(name)
                layout = _resolve_layout(mapping, sheet_mapping, inferred_sheet)
                header_row = _resolve_header_row(mapping, sheet_mapping, inferred_sheet)
                low_confidence = bool(getattr(inferred_sheet, "low_confidence", False))
                if low_confidence:
                    report.low_confidence_sheets.append(name)

                if layout == "wide-table":
                    sheet_recipes, sheet_report = _convert_wide_table(
                        path,
                        name,
                        values_sheet,
                        meta_sheet,
                        merged_map,
                        mapping,
                        sheet_mapping,
                        header_row,
                    )
                elif layout == "template":
                    sheet_recipes, sheet_report = _convert_template(
                        path,
                        name,
                        values_sheet,
                        meta_sheet,
                        merged_map,
                        mapping,
                        sheet_mapping,
                    )
                elif layout == "tall":
                    sheet_recipes, sheet_report = _convert_tall(
                        path,
                        name,
                        values_sheet,
                        meta_sheet,
                        merged_map,
                        mapping,
                        sheet_mapping,
                        header_row,
                    )
                else:
                    sheet_report = ConversionReport()
                    sheet_report.warnings.append(f"Unknown layout for sheet {name}: {layout}")
                    sheet_recipes = []

                recipes.extend(sheet_recipes)
                _merge_report(report, sheet_report)
                report.per_sheet_counts[name] = report.per_sheet_counts.get(name, 0) + len(
                    sheet_recipes
                )

            for recipe in recipes:
                if recipe.confidence is None:
                    recipe.confidence = score_recipe_candidate(recipe)

                if not recipe.identifier:
                    provenance = recipe.provenance or {}
                    sheet_name = str(provenance.get("sheet") or "sheet")
                    row_index = _resolve_row_index(provenance)
                    sheet_slug = _slugify(sheet_name)
                    recipe_id = f"urn:recipeimport:excel:{file_hash}:{sheet_slug}:r{row_index}"
                    recipe.identifier = recipe_id
                    provenance.setdefault("@id", recipe_id)
                    recipe.provenance = provenance

                provenance = recipe.provenance or {}
                original_row = provenance.get("original_row")
                if original_row is not None:
                    raw_artifacts.append(
                        RawArtifact(
                            importer="excel",
                            sourceHash=file_hash,
                            locationId=_raw_location_id(provenance),
                            extension="json",
                            content={
                                "sheet": provenance.get("sheet"),
                                "row_index": provenance.get("row_index"),
                                "headers": provenance.get("original_headers") or [],
                                "row": original_row,
                            },
                        )
                    )

            extracted_rows: list[dict[str, Any]] = []
            for artifact in raw_artifacts:
                if not isinstance(artifact.content, dict):
                    continue
                if "row" not in artifact.content:
                    continue
                extracted_rows.append(
                    {
                        "sheet": artifact.content.get("sheet"),
                        "row_index": artifact.content.get("row_index"),
                        "headers": artifact.content.get("headers") or [],
                        "row": artifact.content.get("row"),
                    }
                )

            if extracted_rows:
                raw_artifacts.append(
                    RawArtifact(
                        importer="excel",
                        sourceHash=file_hash,
                        locationId="full_rows",
                        extension="json",
                        content={"rows": extracted_rows},
                        metadata={"artifact_type": "extracted_rows"},
                    )
                )

            tip_candidates: list[Any] = []
            for recipe in recipes:
                tip_candidates.extend(
                    extract_tip_candidates_from_candidate(recipe, overrides=overrides)
                )

            tips, recipe_specific, not_tips = partition_tip_candidates(tip_candidates)

            report.total_recipes = len(recipes)
            report.total_tips = len(tips)
            report.total_tip_candidates = len(tip_candidates)
            report.total_general_tips = len(tips)
            report.total_recipe_specific_tips = len(recipe_specific)
            report.total_not_tips = len(not_tips)
            report.samples = [
                {"name": recipe.name, "sheet": recipe.provenance.get("sheet")}
                for recipe in recipes[:3]
            ]
            if tips:
                report.tip_samples = [{"text": tip.text[:80]} for tip in tips[:3]]
            return ConversionResult(
                recipes=recipes,
                tips=tips,
                tipCandidates=tip_candidates,
                rawArtifacts=raw_artifacts,
                report=report,
                workbook=path.stem,
                workbookPath=str(path),
            )
        finally:
            wb_values.close()
            wb_meta.close()


registry.register(ExcelImporter())


def _resolve_layout(
    mapping: MappingConfig | None,
    sheet_mapping: SheetMapping | None,
    inferred_sheet: SheetInspection | None,
) -> str:
    if sheet_mapping and sheet_mapping.layout:
        return sheet_mapping.layout
    if mapping and mapping.default_layout:
        return mapping.default_layout
    if inferred_sheet and inferred_sheet.layout:
        return inferred_sheet.layout
    return "wide-table"


def _resolve_header_row(
    mapping: MappingConfig | None,
    sheet_mapping: SheetMapping | None,
    inferred_sheet: SheetInspection | None,
) -> int | None:
    if sheet_mapping and sheet_mapping.header_row:
        return sheet_mapping.header_row
    if inferred_sheet and inferred_sheet.header_row:
        return inferred_sheet.header_row
    return None


def _resolve_sheet_mapping(mapping: MappingConfig, sheet_name: str) -> SheetMapping | None:
    if not mapping.sheets:
        return None
    for sheet in mapping.sheets:
        if sheet.sheet_name and sheet.sheet_name == sheet_name:
            return sheet
    for sheet in mapping.sheets:
        if sheet.name_pattern and re.search(sheet.name_pattern, sheet_name):
            return sheet
    return None


def _merged_cell_map(meta_sheet, values_sheet) -> dict[tuple[int, int], Any]:
    merged_map: dict[tuple[int, int], Any] = {}
    for cell_range in meta_sheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(cell_range))
        top_left = values_sheet.cell(min_row, min_col).value
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                if row == min_row and col == min_col:
                    continue
                merged_map[(row, col)] = top_left
    return merged_map


def _normalize_text(value: str) -> str:
    cleaned = value.replace("\u00a0", " ")
    cleaned = cleaned.replace("\r\n", "\n").replace("\r", "\n")
    cleaned = re.sub(r"[ \t]+", " ", cleaned)
    cleaned = re.sub(r"\n{3,}", "\n\n", cleaned)
    return cleaned.strip()


def _normalize_label(value: str) -> str:
    cleaned = _normalize_text(value).lower()
    cleaned = re.sub(r"\(.*?\)", "", cleaned)
    cleaned = cleaned.strip()
    cleaned = cleaned.rstrip(":")
    return cleaned


def _slugify(value: str) -> str:
    lowered = value.strip().lower()
    slug = _SLUG_RE.sub("_", lowered).strip("_")
    return slug or "unknown"


def _raw_location_id(provenance: dict[str, Any]) -> str:
    sheet_name = str(provenance.get("sheet") or "sheet")
    row_index = _resolve_row_index(provenance)
    return f"{_slugify(sheet_name)}_r{row_index}"


def _resolve_row_index(provenance: dict[str, Any]) -> int:
    for key in ("row_index", "rowIndex", "row"):
        if key in provenance:
            try:
                return int(provenance[key])
            except (TypeError, ValueError):
                return 0
    location = provenance.get("location")
    if isinstance(location, dict):
        for key in ("row_index", "rowIndex", "row", "chunk_index", "chunkIndex", "chunk"):
            if key in location:
                try:
                    return int(location[key])
                except (TypeError, ValueError):
                    return 0
    return 0


def _cell_value(values_sheet, meta_sheet, row: int, col: int, merged_map: dict[tuple[int, int], Any]) -> Any:
    value = values_sheet.cell(row, col).value
    if value is None:
        value = merged_map.get((row, col))
    if value is None:
        cell = meta_sheet.cell(row, col)
        if cell.data_type == "f":
            value = cell.value
    return value


def _row_values(
    values_sheet,
    meta_sheet,
    row: int,
    max_col: int,
    merged_map: dict[tuple[int, int], Any],
) -> list[Any]:
    values: list[Any] = []
    for col in range(1, max_col + 1):
        values.append(_cell_value(values_sheet, meta_sheet, row, col, merged_map))
    return values


def _build_alias_lookup(sheet_mapping: SheetMapping | None) -> dict[str, str]:
    lookup: dict[str, str] = {}
    for field, aliases in _HEADER_ALIASES.items():
        for alias in aliases:
            lookup[_normalize_label(alias)] = field
    if sheet_mapping:
        for field, aliases in sheet_mapping.column_aliases.items():
            for alias in aliases:
                lookup[_normalize_label(str(alias))] = field
    return lookup


def _detect_header_row(
    values_sheet,
    meta_sheet,
    merged_map: dict[tuple[int, int], Any],
    max_rows: int = 30,
) -> tuple[int | None, list[str], float]:
    best_row: int | None = None
    best_score = 0.0
    best_headers: list[str] = []
    max_row = min(values_sheet.max_row or 0, max_rows)
    max_col = values_sheet.max_column or 0
    for row_idx in range(1, max_row + 1):
        row_values = _row_values(values_sheet, meta_sheet, row_idx, max_col, merged_map)
        cleaned = [value for value in row_values if value is not None and str(value).strip()]
        if len(cleaned) < 2:
            continue
        labels = [_normalize_label(str(value)) for value in cleaned]
        unique_ratio = len(set(labels)) / max(len(labels), 1)
        text_ratio = sum(1 for value in cleaned if isinstance(value, str)) / len(cleaned)
        score = len(cleaned) + unique_ratio + text_ratio
        if score > best_score:
            best_score = score
            best_row = row_idx
            best_headers = [
                _normalize_text(str(value)) if value is not None else "" for value in row_values
            ]
    return best_row, best_headers, best_score


def _template_cell_map(
    values_sheet,
    meta_sheet,
    merged_map: dict[tuple[int, int], Any],
) -> tuple[dict[str, str], float]:
    max_row = min(values_sheet.max_row or 0, 40)
    hits: dict[str, str] = {}
    for row_idx in range(1, max_row + 1):
        label_value = _cell_value(values_sheet, meta_sheet, row_idx, 1, merged_map)
        data_value = _cell_value(values_sheet, meta_sheet, row_idx, 2, merged_map)
        if not label_value or data_value is None:
            continue
        label = _normalize_label(str(label_value))
        field = _alias_to_field(label)
        if field and field not in hits:
            hits[field] = f"B{row_idx}"
    score = float(len(hits))
    return hits, score


def _alias_to_field(label: str) -> str | None:
    if not _ALIAS_LOOKUP:
        for field, aliases in _HEADER_ALIASES.items():
            for alias in aliases:
                _ALIAS_LOOKUP[_normalize_label(alias)] = field
    return _ALIAS_LOOKUP.get(label)


def _apply_named_ranges(
    sheet_name: str,
    wb_meta,
    hits: dict[str, str],
) -> None:
    try:
        defined_names = list(wb_meta.defined_names.definedName)
    except AttributeError:
        defined_names = list(wb_meta.defined_names)
    for defined in defined_names:
        name = _normalize_label(defined.name)
        field = _alias_to_field(name)
        if not field:
            continue
        try:
            destinations = list(defined.destinations)
        except Exception:
            destinations = []
        for defined_sheet, coord in destinations:
            if defined_sheet == sheet_name and field not in hits:
                hits[field] = coord


def _analyze_sheet(values_sheet, meta_sheet, merged_map: dict[tuple[int, int], Any]) -> SheetAnalysis:
    header_row, headers, header_score = _detect_header_row(values_sheet, meta_sheet, merged_map)
    alias_lookup = _build_alias_lookup(None)
    matched_headers: dict[str, list[str]] = {}
    for header in headers:
        field = alias_lookup.get(_normalize_label(header))
        if field:
            matched_headers.setdefault(field, []).append(header)

    template_cells, template_score = _template_cell_map(values_sheet, meta_sheet, merged_map)
    _apply_named_ranges(values_sheet.title, meta_sheet.parent, template_cells)

    tall_score = 0.0
    tall_keys: dict[str, str] = {}
    if header_row and headers:
        normalized_headers = [_normalize_label(header) for header in headers]
        recipe_col = _find_column(normalized_headers, ("recipe", "name", "title"))
        section_col = _find_column(normalized_headers, ("section", "type", "field", "kind"))
        value_col = _find_column(normalized_headers, ("text", "value", "line", "item", "content"))
        if recipe_col is not None and section_col is not None and value_col is not None:
            tall_score = header_score + 2.0
            tall_keys = {
                "recipe": headers[recipe_col],
                "section": headers[section_col],
                "value": headers[value_col],
            }

    wide_score = header_score if header_row else 0.0
    layout_scores = {
        "wide-table": wide_score,
        "template": template_score,
        "tall": tall_score,
    }
    layout = max(layout_scores, key=layout_scores.get)
    max_score = layout_scores[layout]
    sorted_scores = sorted(layout_scores.values(), reverse=True)
    ambiguous = len(sorted_scores) > 1 and sorted_scores[0] - sorted_scores[1] < 1.0
    low_confidence = max_score < 2.0 or ambiguous
    if low_confidence:
        layout = "wide-table"

    warnings: list[str] = []
    if low_confidence:
        warnings.append("Low confidence layout detection; defaulted to wide-table.")

    return SheetAnalysis(
        layout=layout,
        header_row=header_row,
        headers=headers,
        column_aliases={key: value for key, value in matched_headers.items() if key},
        template_cells=template_cells,
        tall_keys=tall_keys,
        confidence=max_score,
        low_confidence=low_confidence,
        warnings=warnings,
    )


def _find_column(headers: list[str], aliases: Iterable[str]) -> int | None:
    for idx, header in enumerate(headers):
        if header in aliases:
            return idx
    return None


def _split_lines(value: Any, delimiters: list[str]) -> list[str]:
    if value is None:
        return []
    if isinstance(value, list):
        lines = [str(item) for item in value]
    else:
        text = _normalize_text(str(value))
        for delim in delimiters:
            if delim and delim != "\n":
                text = text.replace(delim, "\n")
        lines = text.split("\n")
    return [line.strip() for line in lines if line and str(line).strip()]


def _normalize_ingredient_line(line: str) -> str:
    return re.sub(r"^\s*[-*]+\s+", "", line).strip()


def _normalize_instruction_line(line: str) -> str:
    cleaned = re.sub(r"^\s*[-*]+\s+", "", line)
    cleaned = re.sub(r"^\s*\d+[.)]\s+", "", cleaned)
    return cleaned.strip()


def _split_ingredients(value: Any, mapping: MappingConfig | None) -> list[str]:
    delimiters = mapping.ingredient_delimiters if mapping else ["\n"]
    lines = _split_lines(value, delimiters)
    return [_normalize_ingredient_line(line) for line in lines if _normalize_ingredient_line(line)]


def _split_instructions(value: Any, mapping: MappingConfig | None) -> list[str]:
    delimiters = mapping.instruction_delimiters if mapping else ["\n"]
    lines = _split_lines(value, delimiters)
    return [
        _normalize_instruction_line(line) for line in lines if _normalize_instruction_line(line)
    ]


def _extract_sections_from_blob(
    text: str,
    mapping: MappingConfig | None,
) -> dict[str, list[str]]:
    """
    Parse a text blob containing section headers into ingredients/instructions/notes.

    Detects headers like:
    - "Ingredients", "INGREDIENTS", "Ingredients:"
    - "Instructions", "Directions", "Method", "Steps"
    - "Notes", "NOTES", "Tips"

    Returns dict with keys: 'ingredients', 'instructions', 'notes'
    If no section headers found, returns empty dict (caller should fall back).
    """
    normalized = _normalize_text(str(text))
    lines = normalized.split("\n")

    sections: dict[str, list[str]] = {
        "ingredients": [],
        "instructions": [],
        "notes": [],
    }
    current_section: str | None = None
    found_any_header = False

    for line in lines:
        stripped = line.strip()
        if not stripped:
            continue

        # Check if this line is a section header
        match = _SECTION_HEADER_RE.match(stripped)
        if match:
            header = match.group(1).lower()
            found_any_header = True
            # Map header to section key
            if header in ("ingredient", "ingredients"):
                current_section = "ingredients"
            elif header in ("instruction", "instructions", "direction", "directions", "method", "step", "steps", "preparation"):
                current_section = "instructions"
            elif header in ("note", "notes", "tip", "tips"):
                current_section = "notes"
            continue

        # Accumulate line under current section
        if current_section:
            if current_section == "ingredients":
                normalized_line = _normalize_ingredient_line(stripped)
                if normalized_line:
                    sections["ingredients"].append(normalized_line)
            elif current_section == "instructions":
                normalized_line = _normalize_instruction_line(stripped)
                if normalized_line:
                    sections["instructions"].append(normalized_line)
            elif current_section == "notes":
                sections["notes"].append(stripped)

    # If no headers were found, return empty dict to signal fallback
    if not found_any_header:
        return {}

    return sections


def _normalize_tags(value: Any) -> list[str]:
    if value is None:
        return []
    if isinstance(value, list):
        raw = value
    else:
        raw = re.split(r",|\n", str(value))
    tags = [_normalize_text(str(tag)) for tag in raw if str(tag).strip()]
    return [tag for tag in tags if tag]


def _build_provenance(
    path: Path,
    sheet_name: str,
    row_index: int,
    headers: list[str] | None,
    row_payload: dict[str, Any] | None,
) -> dict[str, Any]:
    timestamp = dt.datetime.now(dt.timezone.utc).replace(microsecond=0).isoformat()
    if timestamp.endswith("+00:00"):
        timestamp = timestamp[:-6] + "Z"
    return {
        "file_path": str(path),
        "workbook": path.name,
        "sheet": sheet_name,
        "row_index": row_index,
        "original_headers": headers or [],
        "original_row": row_payload or {},
        "import_timestamp": timestamp,
        "converter_version": __version__,
    }


def _convert_wide_table(
    path: Path,
    sheet_name: str,
    values_sheet,
    meta_sheet,
    merged_map: dict[tuple[int, int], Any],
    mapping: MappingConfig | None,
    sheet_mapping: SheetMapping | None,
    header_row: int | None,
) -> tuple[list[RecipeCandidate], ConversionReport]:
    report = ConversionReport()
    if header_row is None:
        report.warnings.append(f"No header row detected for sheet {sheet_name}.")
        return [], report

    max_col = values_sheet.max_column or 0
    headers = _row_values(values_sheet, meta_sheet, header_row, max_col, merged_map)
    header_labels = [
        _normalize_text(str(value)) if value is not None else "" for value in headers
    ]
    alias_lookup = _build_alias_lookup(sheet_mapping)
    field_by_col: dict[int, str] = {}
    for idx, header in enumerate(header_labels):
        label = _normalize_label(header)
        field = alias_lookup.get(label)
        if field:
            field_by_col[idx] = field

    recipes: list[RecipeCandidate] = []
    max_row = values_sheet.max_row or header_row
    for row_idx in range(header_row + 1, max_row + 1):
        if mapping and row_idx in mapping.skip_rows:
            report.skipped_rows.append(
                SkippedRow(sheet=sheet_name, rowIndex=row_idx, reason="Skipped by mapping.")
            )
            continue
        row_values = _row_values(values_sheet, meta_sheet, row_idx, max_col, merged_map)
        if not any(value is not None and str(value).strip() for value in row_values):
            continue

        row_payload: dict[str, Any] = {}
        for idx, value in enumerate(row_values):
            if idx < len(header_labels) and header_labels[idx]:
                row_payload[header_labels[idx]] = _normalize_text(str(value)) if value else ""

        fields: dict[str, Any] = {}
        all_tags: list[str] = []
        for idx, field in field_by_col.items():
            if idx < len(row_values):
                val = row_values[idx]
                if field == "tags":
                    all_tags.extend(_normalize_tags(val))
                else:
                    fields[field] = val
        if all_tags:
            fields["tags"] = all_tags

        name = fields.get("name")
        raw_ingredients = fields.get("ingredients")
        raw_instructions = fields.get("instructions")

        # If we have ingredients but no instructions, check for embedded sections
        if raw_ingredients and not raw_instructions:
            sections = _extract_sections_from_blob(str(raw_ingredients), mapping)
            if sections.get("instructions"):  # Found embedded instructions
                ingredients = sections["ingredients"]
                instructions = sections["instructions"]
                # Handle notes section - append to description if present
                notes_text = "\n".join(sections.get("notes", []))
            else:
                # No section headers found, treat all as ingredients (backward compatible)
                ingredients = _split_ingredients(raw_ingredients, mapping)
                instructions = []
                notes_text = ""
        else:
            ingredients = _split_ingredients(raw_ingredients, mapping)
            instructions = _split_instructions(raw_instructions, mapping)
            notes_text = ""

        description = _normalize_text(str(fields.get("description"))) if fields.get("description") else None
        # Append notes to description if found from section parsing
        if notes_text:
            if description:
                description = f"{description}\n\n{notes_text}"
            else:
                description = notes_text
        recipe_yield = (
            _normalize_text(str(fields.get("recipeYield"))) if fields.get("recipeYield") else None
        )
        source_url = _normalize_text(str(fields.get("sourceUrl"))) if fields.get("sourceUrl") else None
        tags = _normalize_tags(fields.get("tags"))

        if not name or not str(name).strip():
            report.skipped_rows.append(
                SkippedRow(sheet=sheet_name, rowIndex=row_idx, reason="Missing name.")
            )
            report.missing_field_counts["name"] = report.missing_field_counts.get("name", 0) + 1
            continue

        if not ingredients:
            report.missing_field_counts["ingredients"] = (
                report.missing_field_counts.get("ingredients", 0) + 1
            )
        if not instructions:
            report.missing_field_counts["instructions"] = (
                report.missing_field_counts.get("instructions", 0) + 1
            )

        provenance = _build_provenance(path, sheet_name, row_idx, header_labels, row_payload)
        recipe = RecipeCandidate(
            name=str(name),
            ingredients=ingredients,
            instructions=instructions,
            description=description,
            recipeYield=recipe_yield,
            sourceUrl=source_url,
            tags=tags,
            provenance=provenance,
        )
        recipes.append(recipe)

    return recipes, report


def _convert_template(
    path: Path,
    sheet_name: str,
    values_sheet,
    meta_sheet,
    merged_map: dict[tuple[int, int], Any],
    mapping: MappingConfig | None,
    sheet_mapping: SheetMapping | None,
) -> tuple[list[RecipeCandidate], ConversionReport]:
    report = ConversionReport()
    template_cells: dict[str, str] = {}
    if sheet_mapping:
        template_cells.update(sheet_mapping.template_cells)
    if not template_cells:
        template_cells, _ = _template_cell_map(values_sheet, meta_sheet, merged_map)
        _apply_named_ranges(sheet_name, meta_sheet.parent, template_cells)

    if not template_cells:
        report.warnings.append(f"No template cells detected for sheet {sheet_name}.")
        return [], report

    values: dict[str, Any] = {}
    for field, coord in template_cells.items():
        cell = values_sheet[coord]
        value = cell.value
        if value is None:
            value = merged_map.get((cell.row, cell.column))
        if value is None:
            meta_cell = meta_sheet[coord]
            if meta_cell.data_type == "f":
                value = meta_cell.value
        values[field] = value

    name = values.get("name")
    ingredients = _split_ingredients(values.get("ingredients"), mapping)
    instructions = _split_instructions(values.get("instructions"), mapping)
    description = _normalize_text(str(values.get("description"))) if values.get("description") else None
    recipe_yield = (
        _normalize_text(str(values.get("recipeYield"))) if values.get("recipeYield") else None
    )
    source_url = _normalize_text(str(values.get("sourceUrl"))) if values.get("sourceUrl") else None
    tags = _normalize_tags(values.get("tags"))

    if not name or not str(name).strip():
        report.missing_field_counts["name"] = report.missing_field_counts.get("name", 0) + 1
        report.warnings.append(f"Missing name for template sheet {sheet_name}.")
        return [], report

    if not ingredients:
        report.missing_field_counts["ingredients"] = report.missing_field_counts.get("ingredients", 0) + 1
    if not instructions:
        report.missing_field_counts["instructions"] = report.missing_field_counts.get("instructions", 0) + 1

    provenance = _build_provenance(path, sheet_name, 1, list(template_cells.keys()), values)
    recipe = RecipeCandidate(
        name=str(name),
        ingredients=ingredients,
        instructions=instructions,
        description=description,
        recipeYield=recipe_yield,
        sourceUrl=source_url,
        tags=tags,
        provenance=provenance,
    )
    return [recipe], report


def _convert_tall(
    path: Path,
    sheet_name: str,
    values_sheet,
    meta_sheet,
    merged_map: dict[tuple[int, int], Any],
    mapping: MappingConfig | None,
    sheet_mapping: SheetMapping | None,
    header_row: int | None,
) -> tuple[list[RecipeCandidate], ConversionReport]:
    report = ConversionReport()
    if header_row is None:
        report.warnings.append(f"No header row detected for tall sheet {sheet_name}.")
        return [], report

    max_col = values_sheet.max_column or 0
    headers = _row_values(values_sheet, meta_sheet, header_row, max_col, merged_map)
    header_labels = [
        _normalize_text(str(value)) if value is not None else "" for value in headers
    ]
    normalized_headers = [_normalize_label(label) for label in header_labels]

    recipe_header = "recipe"
    section_header = "section"
    value_header = "value"
    if sheet_mapping and sheet_mapping.tall_keys:
        recipe_header = sheet_mapping.tall_keys.get("recipe", recipe_header)
        section_header = sheet_mapping.tall_keys.get("section", section_header)
        value_header = sheet_mapping.tall_keys.get("value", value_header)

    recipe_col = _find_column(normalized_headers, (_normalize_label(recipe_header), "recipe", "name", "title"))
    section_col = _find_column(normalized_headers, (_normalize_label(section_header), "section", "type", "field"))
    value_col = _find_column(normalized_headers, (_normalize_label(value_header), "value", "text", "line"))

    if recipe_col is None or section_col is None or value_col is None:
        report.warnings.append(f"Missing tall layout columns for sheet {sheet_name}.")
        return [], report

    recipes: dict[str, dict[str, Any]] = {}
    max_row = values_sheet.max_row or header_row
    current_recipe = ""
    for row_idx in range(header_row + 1, max_row + 1):
        row_values = _row_values(values_sheet, meta_sheet, row_idx, max_col, merged_map)
        recipe_name = row_values[recipe_col] if recipe_col < len(row_values) else None
        section = row_values[section_col] if section_col < len(row_values) else None
        value = row_values[value_col] if value_col < len(row_values) else None
        if recipe_name is None and current_recipe:
            recipe_name = current_recipe
        if not recipe_name:
            continue
        current_recipe = str(recipe_name)

        recipe_entry = recipes.setdefault(
            current_recipe,
            {
                "ingredients": [],
                "instructions": [],
                "first_row": row_idx,
                "original_rows": [],
            },
        )
        recipe_entry["original_rows"].append(
            {
                "row_index": row_idx,
                header_labels[recipe_col]: _normalize_text(str(recipe_name)),
                header_labels[section_col]: _normalize_text(str(section)) if section else "",
                header_labels[value_col]: _normalize_text(str(value)) if value else "",
            }
        )
        if not value:
            continue
        section_label = _normalize_label(str(section)) if section else ""
        if any(token in section_label for token in _SECTION_INSTRUCTIONS):
            recipe_entry["instructions"].extend(_split_instructions(value, mapping))
        elif any(token in section_label for token in _SECTION_INGREDIENTS):
            recipe_entry["ingredients"].extend(_split_ingredients(value, mapping))
        else:
            recipe_entry["ingredients"].extend(_split_ingredients(value, mapping))

    recipe_candidates: list[RecipeCandidate] = []
    for name, data in recipes.items():
        ingredients = data.get("ingredients", [])
        instructions = data.get("instructions", [])
        if not ingredients:
            report.missing_field_counts["ingredients"] = report.missing_field_counts.get("ingredients", 0) + 1
        if not instructions:
            report.missing_field_counts["instructions"] = report.missing_field_counts.get("instructions", 0) + 1
        provenance = _build_provenance(
            path,
            sheet_name,
            data.get("first_row", 1),
            header_labels,
            {"rows": data.get("original_rows", [])},
        )
        recipe_candidates.append(
            RecipeCandidate(
                name=str(name),
                ingredients=ingredients,
                instructions=instructions,
                provenance=provenance,
            )
        )

    return recipe_candidates, report


def _merge_report(target: ConversionReport, source: ConversionReport) -> None:
    target.skipped_rows.extend(source.skipped_rows)
    target.warnings.extend(source.warnings)
    target.errors.extend(source.errors)
    for key, value in source.missing_field_counts.items():
        target.missing_field_counts[key] = target.missing_field_counts.get(key, 0) + value
