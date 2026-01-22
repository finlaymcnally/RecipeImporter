from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook


FIXTURES_DIR = Path(__file__).resolve().parent


def _write_wide_table(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Title", "Ingredients", "Instructions", "Yield", "Tags"])
    ws.append(
        [
            "  Lemon Bread  ",
            "1 cup flour\n2 eggs",
            "1. Mix\n2. Bake",
            "1 loaf",
            "baking, dessert",
        ]
    )
    ws.append(
        [
            "Toast",
            "- 2 slices bread\n- butter",
            "- Toast bread\n- Spread butter",
            "2 servings",
            "breakfast",
        ]
    )

    ws2 = wb.create_sheet("Formula")
    ws2.append(["Title", "Ingredients", "Instructions", "Description"])
    ws2.append(["Formula Recipe", "1 cup sugar", "Mix", None])
    ws2["D2"] = '="Formula note"'

    wb.save(path)


def _write_template(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Template"
    ws["A1"] = "Title"
    ws["B1"] = "  Template Soup  "
    ws["A2"] = "Ingredients"
    ws["B2"] = "1 cup water\nSalt"
    ws["A3"] = "Instructions"
    ws["B3"] = "Boil\nServe"
    ws["A4"] = "Yield"
    ws["B4"] = "4 servings"
    ws["A5"] = "Notes"
    ws["B5"] = "A simple soup"

    wb.save(path)


def _write_tall_relational(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "TallSheet"
    ws.append(["Recipe", "Section", "Text"])
    ws.append(["Cereal", "Ingredient", "1 cup cereal"])
    ws.append([None, "Ingredient", "1 cup milk"])
    ws.append([None, "Instruction", "Combine and eat"])
    ws.merge_cells("A2:A4")
    ws.append(["Salad", "Ingredient", "Lettuce"])
    ws.append([None, "Instruction", "Toss"])
    ws.merge_cells("A5:A6")

    wb.save(path)


def main() -> None:
    FIXTURES_DIR.mkdir(parents=True, exist_ok=True)
    _write_wide_table(FIXTURES_DIR / "wide_table.xlsx")
    _write_template(FIXTURES_DIR / "template.xlsx")
    _write_tall_relational(FIXTURES_DIR / "tall_relational.xlsx")


if __name__ == "__main__":
    main()
